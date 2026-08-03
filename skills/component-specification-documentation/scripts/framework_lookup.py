"""
Read the official TM Forum framework releases: descriptions from the spreadsheets, names/tokens from the
JSON definitions.

Shared by fill_descriptions_from_frameworks.py and align_framework_names.py.

The frameworks folder is **not** part of this repository -- these are published TM Forum releases, not
component data -- so its location is always passed in. Ask the user for it rather than guessing: a stale
hard-coded path fails silently, making descriptions look genuinely absent so a run emits
"(no description available)" instead of stopping to ask.

Two rules govern which file is read:

* **One release per entry, not one release per run.** Every `componentMetadata` entry carries its own
  version suffix and a component's list is routinely mixed -- TMFC003 has 39 Functional Framework entries
  at v25.5 and 2 at v23.0. The same id can exist in several releases with different wording (function 208
  does), so a single-workbook lookup silently returns the wrong revision.
* **Filenames vary between releases.** `GB1033F_Functional_Framework_Excel_v25.5.xlsx`,
  `GB1033F_..._Excel_Format_v24.5.xlsx` and `GB1033_..._Excel_v23.0.0.xlsx` are all the same family, and a
  YAML version of `v23.0` must match a file named `v23.0.0`. Resolution is therefore by family pattern
  plus version prefix, not by exact filename.
"""
import glob
import html
import os
import re

FAMILIES = {
    # key: (glob pattern, description)
    "ff_xlsx":   ("GB1033*Functional_Framework*.xlsx", "Functional Framework spreadsheet"),
    "etom_xlsx": ("GB921*Business_Process_Framework*.xlsx", "eTOM spreadsheet"),
    "sid_xlsx":  ("GB922*Information_Framework*.xlsx", "SID spreadsheet"),
    "ff_json":   ("functionalFramework_v*.json", "Functional Framework definitions"),
    "etom_json": ("etom_v*.json", "eTOM definitions"),
    "sid_json":  ("sid_v*.json", "SID definitions"),
}

_VER_IN_NAME = re.compile(r"v(\d+(?:\.\d+)*)", re.IGNORECASE)


def clean(text):
    """Flowing prose from a spreadsheet cell.

    De-hyphenation runs before the newline collapse: a word broken across lines ("pre-\\norder") must
    rejoin without a space, or it reads as "pre- order". Some framework descriptions also carry raw HTML
    rather than prose -- Functional Framework function 205 holds `&amp;` and a `<ul><li>` list -- which
    would otherwise be written into a Markdown table and rendered as literal markup."""
    if text is None:
        return ""
    text = str(text)
    text = re.sub(r"(?i)</\s*li\s*>", " ", text)
    text = re.sub(r"(?i)<\s*li\s*>", " - ", text)
    text = re.sub(r"<[^>]{1,80}>", " ", text)
    text = html.unescape(text)
    for bad, good in (("•", "- "), ("“", '"'), ("”", '"'),
                      ("‘", "'"), ("’", "'"), ("–", "-"), (" ", " ")):
        text = text.replace(bad, good)
    text = re.sub(r"(\w)-\s*\n\s*(\w)", r"\1-\2", text)
    text = re.sub(r"\s*\n\s*", " ", text)
    return re.sub(r"\s{2,}", " ", text).strip()


def escape_cell(text):
    """An unescaped pipe inside a cell silently creates phantom columns in most Markdown renderers."""
    return (text or "").replace("|", r"\|")


def norm_version(v):
    return str(v or "").strip().lstrip("vV")


def resolve(frameworks_dir, family, version):
    """Path to the file of `family` for release `version`.

    A YAML version like `v23.0` legitimately matches a file named `v23.0.0` or `v23.0.1`, so an exact
    match is preferred and a prefix match accepted; among several prefix matches the shortest (closest)
    version wins."""
    if family not in FAMILIES:
        raise ValueError(f"unknown framework family {family!r}")
    pattern, label = FAMILIES[family]
    candidates = glob.glob(os.path.join(frameworks_dir, pattern))
    if not candidates:
        raise FileNotFoundError(
            f"no {label} found in {frameworks_dir}\n"
            f"  expected something matching {pattern}\n"
            f"  ask the user where the framework releases are kept")

    wanted = norm_version(version)
    scored = []
    for path in candidates:
        m = _VER_IN_NAME.search(os.path.basename(path))
        if not m:
            continue
        found = m.group(1)
        if found == wanted:
            scored.append((0, len(found), path))
        elif found.startswith(wanted + ".") or wanted.startswith(found + "."):
            scored.append((1, len(found), path))
    if not scored:
        have = sorted({(_VER_IN_NAME.search(os.path.basename(p)) or [None, "?"])[1]
                       for p in candidates})
        raise FileNotFoundError(
            f"no {label} for version {version!r} in {frameworks_dir}; available: {', '.join(have)}")
    scored.sort()
    return scored[0][2]


# ---------------------------------------------------------------- JSON definitions (names / tokens)

def load_definitions(frameworks_dir, family, version):
    """id -> entry dict from a framework JSON release.

    Each entry carries `name` and `token`, where `token` is exactly the underscore form the component
    YAML's pipe-delimited entries use (`Customer_Order_Change_Management`). That makes name alignment a
    direct token-to-token comparison rather than prose matching."""
    import json
    path = resolve(frameworks_dir, family, version)
    with open(path, encoding="utf-8") as f:
        doc = json.load(f)
    entries = doc.get("entries", doc if isinstance(doc, list) else [])
    index, conflicts = {}, []
    for e in entries:
        key = str(e.get("id", "")).strip()
        if not key:
            continue
        if key in index:
            prior = index[key]
            if (prior.get("token"), prior.get("level")) != (e.get("token"), e.get("level")):
                conflicts.append(key)
            continue          # duplicates that agree are common (one row per domain); take the first
        index[key] = e
    return index, os.path.basename(path), sorted(set(conflicts))


# ---------------------------------------------------------------- spreadsheets (descriptions)

_HEADER_MARKERS = ("function id", "function name", "function description", "aggregate function",
                   "process identifier", "brief description", "extended description", "abe name")


def _norm_header(text):
    """Trailing apostrophes are real: one Functional Framework sheet spells a header `AF Lev.2'`."""
    return re.sub(r"\s+", " ", str(text or "").strip().rstrip("'")).lower()


def _pick_column(headers, labels):
    """Exact normalized match only.

    Deliberately not a prefix or substring match: the Functional Framework sheet puts `AF Lev.1` (the
    aggregate function *name*) directly beside `AF L1` (a hierarchy id like `6.7.1`), and a loose match
    silently reads the wrong column."""
    for label in labels:
        target = _norm_header(label)
        for i, h in enumerate(headers):
            if _norm_header(h) == target:
                return i
    return None


def _sheet_rows(workbook, sheet_name):
    return list(workbook[sheet_name].iter_rows(values_only=True))


def load_ff_descriptions(frameworks_dir, version):
    """function id -> {description, af1, af2} for one Functional Framework release."""
    import openpyxl
    path = resolve(frameworks_dir, "ff_xlsx", version)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # 'Functions and AFs' has one fully-populated row per function; the plain 'Functions' sheet has the
    # aggregate-function columns merged, so they read as None on most rows.
    for sheet in ("Functions and AFs", "Functions"):
        if sheet not in wb.sheetnames:
            continue
        rows = _sheet_rows(wb, sheet)
        if not rows:
            continue
        headers = rows[0]
        ci = {
            "id": _pick_column(headers, ["Function ID"]),
            "name": _pick_column(headers, ["Function Name"]),
            "description": _pick_column(headers, ["Description", "Function Description"]),
            "af1": _pick_column(headers, ["AF Lev.1"]),
            "af2": _pick_column(headers, ["AF Lev.2"]),
        }
        if ci["id"] is None or ci["description"] is None:
            continue
        out, last_af1, last_af2 = {}, "", ""
        for row in rows[1:]:
            def cell(key):
                i = ci[key]
                return row[i] if i is not None and i < len(row) else None
            af1, af2 = clean(cell("af1")), clean(cell("af2"))
            # merged aggregate-function cells read as None on all but the first row of the merge
            last_af1, last_af2 = af1 or last_af1, af2 or last_af2
            raw = cell("id")
            if raw is None or not str(raw).strip():
                continue
            key = str(int(raw)) if isinstance(raw, float) and raw.is_integer() else str(raw).strip()
            out[key] = {"name": clean(cell("name")), "description": clean(cell("description")),
                        "af1": last_af1, "af2": last_af2}
        if out:
            return out, os.path.basename(path), sheet
    raise RuntimeError(f"no usable functions sheet in {path}")


def load_etom_descriptions(frameworks_dir, version):
    """process identifier -> {name, description} for one eTOM release.

    The functions sheet is version-named with a comma (`eTOM25,5`), so it's found by prefix while
    excluding `eTOM Deleted`. `Brief description` is used rather than `Extended Description`: it matches
    the published component documents' own 2.1 prose almost verbatim, where the extended variant is much
    longer. Identifiers repeat (one row per domain / vertical group); the first is taken."""
    import openpyxl
    path = resolve(frameworks_dir, "etom_xlsx", version)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = [s for s in wb.sheetnames
              if s.lower().startswith("etom") and "deleted" not in s.lower()]
    for sheet in sheets:
        rows = _sheet_rows(wb, sheet)
        if not rows:
            continue
        headers = rows[0]
        ci = {
            "id": _pick_column(headers, ["Process identifier"]),
            "name": _pick_column(headers, ["Process"]),
            "description": _pick_column(headers, ["Brief description"]),
            "extended": _pick_column(headers, ["Extended Description"]),
        }
        if ci["id"] is None or ci["description"] is None:
            continue
        out = {}
        for row in rows[1:]:
            def cell(key):
                i = ci[key]
                return row[i] if i is not None and i < len(row) else None
            raw = cell("id")
            if raw is None or not str(raw).strip():
                continue
            key = str(raw).strip()
            if key in out:
                continue
            desc = clean(cell("description")) or clean(cell("extended"))
            out[key] = {"name": clean(cell("name")), "description": desc}
        if out:
            return out, os.path.basename(path), sheet
    raise RuntimeError(f"no usable eTOM sheet in {path} (sheets: {wb.sheetnames})")


def load_sid_abe_descriptions(frameworks_dir, version):
    """ABE name -> definition for one SID release.

    ABE-level rows are the ones with a dot-qualified `ABE Name` and empty `BE Name`/`Attribute Name`.
    Only the first sentence is kept: the full `Documentation` runs 700-900 characters, which swamps a
    table column, while the first sentence is a complete self-contained definition."""
    import openpyxl
    path = resolve(frameworks_dir, "sid_xlsx", version)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = "All Domains" if "All Domains" in wb.sheetnames else wb.sheetnames[0]
    rows = _sheet_rows(wb, sheet)
    headers = rows[0] if rows else []
    ci = {k: _pick_column(headers, [v]) for k, v in
          (("abe", "ABE Name"), ("be", "BE Name"), ("attr", "Attribute Name"),
           ("doc", "Documentation"))}
    if ci["abe"] is None or ci["doc"] is None:
        raise RuntimeError(f"unexpected SID sheet layout in {path}")
    out = {}
    for row in rows[1:]:
        def cell(key):
            i = ci[key]
            return row[i] if i is not None and i < len(row) else None
        abe, be, attr = clean(cell("abe")), clean(cell("be")), clean(cell("attr"))
        if not abe or be or attr:
            continue
        leaf = abe.split(".")[-1].strip()
        if leaf in out:
            continue
        doc = clean(cell("doc"))
        if doc:
            first = doc.split(". ")[0].rstrip(".") + "."
            out[leaf] = first
    return out, os.path.basename(path), sheet
